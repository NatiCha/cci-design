"""
Invoice Generation Script (Excel Version)

Generates client invoices from monthly timesheet reports. Reads timesheet data
from an Apple Numbers file, uses an Excel template to create professionally formatted
invoice cover sheets per project with Excel formulas, and attaches detailed timesheet breakdowns.

Usage:
    uv run python print_sheets/split_by_project.py <input_file.numbers>

Example:
    uv run python print_sheets/split_by_project.py data/timesheet_monthly_report_2025_11_a.numbers
"""

import argparse
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from numbers_parser import Document
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Border, Font, PatternFill, Color

# Import constants from shared library
sys.path.insert(0, str(Path(__file__).parent.parent))
from timesheet_lib import NON_PROJECT_NAMES

# =============================================================================
# CONSTANTS
# =============================================================================

TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "invoice-template.xlsx"
LOGO_PATH = Path(__file__).parent.parent / "templates" / "cci-design-logo.png"

# Phase code to template description mapping
PHASE_TO_DESCRIPTION = {
    "PD": "Pre-Design",
    "SD": "Schematic Design",
    "DD": "Design Development",
    "CD": "Construction Documents",
    "CA": "Construction Administration",
    "M": "Meetings w/ Client or Contractor",
}

# Task code to template description mapping
TASK_TO_DESCRIPTION = {
    "DP": "Design Principal",
    "PM": "Project Management",
    "3-D": "3D Model",
    "D-D": "Design and Documentation",
    "M": "Meetings",
}

# Processing order
PHASE_ORDER = ["PD", "SD", "DD", "CD", "CA", "M"]
TASK_ORDER_STANDARD = ["DP", "PM", "3-D", "D-D"]
TASK_ORDER_MEETINGS = ["M"]

# Valid codes for billable entries (excludes BD, NA which are for non-projects)
BILLABLE_TASK_CODES = {"DP", "PM", "3-D", "D-D", "M"}
BILLABLE_PHASE_CODES = {"PD", "SD", "DD", "CD", "CA", "M"}

# Template row indices (1-indexed for Excel)
# Based on actual template inspection
TEMPLATE_STRUCTURE = {
    "project_name_row": 3,
    "project_number_row": 4,
    "invoice_date_row": 5,
    "invoice_number_row": 6,
    "value_col": 3,  # Column C for placeholder values
    "units_col": 3,  # Column C for Units
    "rate_col": 4,   # Column D for Rate
    "cost_col": 5,   # Column E for Cost
    "total_col": 6,  # Column F for Total
    "phases": {
        "PD": {"header": 10, "tasks": [11, 12, 13, 14], "subtotal": 15},
        "SD": {"header": 16, "tasks": [17, 18, 19, 20], "subtotal": 21},
        "DD": {"header": 22, "tasks": [23, 24, 25, 26], "subtotal": 27},
        "CD": {"header": 28, "tasks": [29, 30, 31, 32], "subtotal": 33},
        "CA": {"header": 34, "tasks": [35, 36, 37, 38], "subtotal": 39},
        "M": {"header": 40, "tasks": [41], "subtotal": 42},
    },
    "overall_subtotal_row": 43,
    "reimbursable_header": 44,
    "reimbursable_rows": [45, 46, 47],
    "reimbursable_subtotal": 48,
    "total_amount_due_row": 50,
    "footer_row": 52,
}

# Task code to row offset within phase section (for standard phases)
TASK_TO_ROW_OFFSET = {"DP": 0, "PM": 1, "3-D": 2, "D-D": 3}

# Excel sheet name limit
MAX_SHEET_NAME_LENGTH = 31


# =============================================================================
# INPUT READING
# =============================================================================


def read_timesheet_data(input_file: Path) -> list[dict]:
    """
    Read timesheet entries from the input Numbers file.

    Reads from sheet "3 Timesheet Detail (Edit)", table "Timesheet Detail (Edit)".

    Returns:
        List of entry dicts with keys: project_id, date, employee, hours, task, phase, wid
    """
    print(f"Reading input file: {input_file}")

    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    doc = Document(str(input_file))

    sheet_name = "3 Timesheet Detail (Edit)"
    table_name = "Timesheet Detail (Edit)"

    try:
        sheet = doc.sheets[sheet_name]
        table = sheet.tables[table_name]
    except (KeyError, IndexError) as e:
        raise ValueError(
            f"Expected sheet '{sheet_name}' with table '{table_name}' not found in input file"
        ) from e

    print(f"Parsing timesheet data from '{sheet_name}'...")

    rows = table.rows(values_only=True)
    if len(rows) < 2:
        raise ValueError("Input file contains no data rows")

    # Expected columns: Project ID, Date, Employee, Hours, Task, Phase, WID
    entries = []
    for row in rows[1:]:  # Skip header
        if len(row) < 7:
            continue

        project_id = str(row[0]).strip() if row[0] else ""
        date_val = row[1]
        employee = str(row[2]).strip() if row[2] else ""
        hours = float(row[3]) if row[3] else 0.0
        task = str(row[4]).strip().upper() if row[4] else ""
        phase = str(row[5]).strip().upper() if row[5] else ""
        wid = str(row[6]).strip() if row[6] else ""

        # Convert date to date object if needed
        if isinstance(date_val, datetime):
            date_val = date_val.date()
        elif isinstance(date_val, str):
            try:
                date_val = datetime.strptime(date_val, "%m/%d/%Y").date()
            except ValueError:
                date_val = None

        entries.append(
            {
                "project_id": project_id,
                "date": date_val,
                "employee": employee,
                "hours": hours,
                "task": task,
                "phase": phase,
                "wid": wid,
            }
        )

    return entries


def parse_project_id(project_id: str) -> tuple[str, str]:
    """
    Parse project ID into name and number.

    Args:
        project_id: String in format "Name: Number"

    Returns:
        Tuple of (project_name, project_number)

    Raises:
        ValueError: If project_id doesn't contain colon separator
    """
    if ":" not in project_id:
        raise ValueError(f"Project ID missing colon separator: '{project_id}'")

    parts = project_id.split(":", 1)
    name = parts[0].strip()
    number = parts[1].strip()
    return name, number


# =============================================================================
# FILTERING & VALIDATION
# =============================================================================


def filter_non_projects(entries: list[dict]) -> tuple[list[dict], int]:
    """
    Remove non-project entries (Office, Vacation, Holiday, Sick, Personal Time).

    Returns:
        Tuple of (filtered entries, count of excluded entries)
    """
    filtered = []
    excluded_count = 0

    for entry in entries:
        project_id = entry["project_id"].lower()
        is_non_project = False

        for non_project in NON_PROJECT_NAMES:
            if project_id.startswith(non_project + ":") or project_id == non_project:
                is_non_project = True
                break

        if is_non_project:
            excluded_count += 1
        else:
            filtered.append(entry)

    return filtered, excluded_count


def filter_zero_hours(entries: list[dict]) -> list[dict]:
    """Remove entries with zero or negative hours."""
    return [e for e in entries if e["hours"] > 0]


def validate_codes(entries: list[dict]) -> list[str]:
    """
    Validate Task and Phase codes against allowed billable codes.

    Returns:
        List of error messages (empty if all valid)
    """
    errors = []

    for entry in entries:
        task = entry["task"]
        phase = entry["phase"]
        project_id = entry["project_id"]

        if task and task not in BILLABLE_TASK_CODES:
            errors.append(
                f"Invalid Task code '{task}' for project '{project_id}' "
                f"(valid: {', '.join(sorted(BILLABLE_TASK_CODES))})"
            )

        if phase and phase not in BILLABLE_PHASE_CODES:
            errors.append(
                f"Invalid Phase code '{phase}' for project '{project_id}' "
                f"(valid: {', '.join(sorted(BILLABLE_PHASE_CODES))})"
            )

    return errors


def validate_project_ids(entries: list[dict]) -> list[str]:
    """
    Validate all Project IDs can be parsed (contain colon).

    Returns:
        List of error messages for unparseable project IDs
    """
    errors = []
    seen = set()

    for entry in entries:
        project_id = entry["project_id"]
        if project_id in seen:
            continue
        seen.add(project_id)

        if ":" not in project_id:
            errors.append(f"Project ID missing colon separator: '{project_id}'")

    return errors


# =============================================================================
# AGGREGATION
# =============================================================================


def aggregate_hours(entries: list[dict]) -> dict[str, dict[tuple[str, str], float]]:
    """
    Aggregate hours by project, then by (task, phase) combination.

    Returns:
        Nested dict: {project_id: {(task_code, phase_code): total_hours}}
    """
    result = defaultdict(lambda: defaultdict(float))

    for entry in entries:
        project_id = entry["project_id"]
        task = entry["task"]
        phase = entry["phase"]
        hours = entry["hours"]

        if task and phase:
            result[project_id][(task, phase)] += hours

    return result


def group_by_project(entries: list[dict]) -> dict[str, list[dict]]:
    """
    Group individual entries by project for detail tables.

    Returns:
        Dict mapping project_id to list of entries, sorted by date
    """
    grouped = defaultdict(list)

    for entry in entries:
        grouped[entry["project_id"]].append(entry)

    # Sort each project's entries by date
    for project_id in grouped:
        grouped[project_id].sort(key=lambda e: e["date"] or date.min)

    return grouped


# =============================================================================
# OUTPUT GENERATION
# =============================================================================


def generate_output_filename(input_file: Path, entries: list[dict]) -> Path:
    """
    Generate output filename with versioning.

    Format: invoices_{month_name}_{year}_a.xlsx
    Adds _a, _b, _c suffix if file exists.
    """
    # Extract month/year from entries
    for entry in entries:
        if entry["date"]:
            month_name = entry["date"].strftime("%B").lower()
            year = entry["date"].year
            break
    else:
        # Fallback to current date
        today = date.today()
        month_name = today.strftime("%B").lower()
        year = today.year

    output_dir = input_file.parent
    base_name = f"invoices_{month_name}_{year}"

    # Check for existing files and add suffix (always start with _a per project convention)
    suffix_char = ord("a")

    while True:
        suffix = f"_{chr(suffix_char)}"
        filename = f"{base_name}{suffix}.xlsx"
        output_path = output_dir / filename
        if not output_path.exists():
            return output_path
        suffix_char += 1
        if suffix_char > ord("z"):
            raise RuntimeError("Too many output files exist")


def ordinal_suffix(day: int) -> str:
    """Return ordinal suffix for a day number (1st, 2nd, 3rd, etc.)."""
    if 11 <= day <= 13:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")


def get_invoice_date() -> str:
    """Get current date formatted as 'December 30, 2025' in EST/EDT."""
    tz = ZoneInfo("America/New_York")
    today = datetime.now(tz).date()
    return today.strftime("%B %d, %Y").replace(" 0", " ")  # Remove leading zero from day


def format_timesheet_date(d: date) -> str:
    """Format date as 'Thursday, October 30th 2025'."""
    day_suffix = ordinal_suffix(d.day)
    # %A = full weekday, %B = full month, %d = day, %Y = year
    base = d.strftime(f"%A, %B {d.day}{day_suffix} %Y")
    return base


def make_sheet_name(name: str, suffix: str) -> str:
    """Create a valid Excel sheet name with suffix (31 chars max).

    Args:
        name: The base name (typically project_id)
        suffix: Suffix to append (e.g., " A" for invoice, " B" for timesheet)

    Returns:
        A valid Excel sheet name (<= 31 chars)
    """
    # Replace colon with space-hyphen-space for readability
    sanitized = name.replace(':', ' -')

    # Replace other invalid characters with hyphen
    invalid_chars = ['\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '-')

    # Calculate max length for the base name (reserve space for suffix)
    max_base_len = MAX_SHEET_NAME_LENGTH - len(suffix)

    # Truncate base name if needed
    if len(sanitized) > max_base_len:
        sanitized = sanitized[:max_base_len].rstrip()

    return sanitized + suffix


# =============================================================================
# EXCEL MANIPULATION
# =============================================================================


def calculate_phase_hours(
    hours_dict: dict[tuple[str, str], float]
) -> dict[str, float]:
    """Calculate total hours per phase."""
    phase_totals = defaultdict(float)
    for (_, phase), hours in hours_dict.items():
        phase_totals[phase] += hours
    return phase_totals


def calculate_rows_to_delete(
    hours_dict: dict[tuple[str, str], float]
) -> list[int]:
    """
    Calculate which rows to delete, returned in descending order.

    Deletes:
    - Entire phase sections if phase has 0 hours
    - Individual task rows with 0 hours in phases that have some hours
    """
    rows_to_delete = []
    phase_totals = calculate_phase_hours(hours_dict)

    for phase_code in PHASE_ORDER:
        phase_info = TEMPLATE_STRUCTURE["phases"][phase_code]
        phase_hours = phase_totals.get(phase_code, 0)

        if phase_hours == 0:
            # Delete entire phase section
            rows_to_delete.append(phase_info["header"])
            rows_to_delete.extend(phase_info["tasks"])
            rows_to_delete.append(phase_info["subtotal"])
        elif phase_code != "M":
            # For non-Meetings phases, delete individual zero-hour task rows
            # Meetings phase only has one task row that aggregates all M phase hours
            for i, task_code in enumerate(TASK_ORDER_STANDARD):
                task_hours = hours_dict.get((task_code, phase_code), 0)
                if task_hours == 0:
                    rows_to_delete.append(phase_info["tasks"][i])

    return sorted(rows_to_delete, reverse=True)


def populate_invoice_sheet(
    ws,
    project_name: str,
    project_number: str,
    hours_dict: dict[tuple[str, str], float],
) -> None:
    """Populate the invoice worksheet with project data."""
    value_col = TEMPLATE_STRUCTURE["value_col"]
    units_col = TEMPLATE_STRUCTURE["units_col"]

    # Fill header fields
    ws.cell(row=TEMPLATE_STRUCTURE["project_name_row"], column=value_col, value=project_name)
    ws.cell(row=TEMPLATE_STRUCTURE["project_number_row"], column=value_col, value=project_number)
    ws.cell(row=TEMPLATE_STRUCTURE["invoice_date_row"], column=value_col, value=get_invoice_date())
    ws.cell(row=TEMPLATE_STRUCTURE["invoice_number_row"], column=value_col, value=f"{project_number} INV ##")

    # Fill hours for each phase/task
    for phase_code in PHASE_ORDER:
        phase_info = TEMPLATE_STRUCTURE["phases"][phase_code]

        if phase_code == "M":
            # For Meetings phase, aggregate ALL tasks into the single Meetings row
            # This handles cases where PM, D-D, or other tasks occur during meetings
            meetings_hours = sum(
                hours for (task, phase), hours in hours_dict.items()
                if phase == "M"
            )
            task_row = phase_info["tasks"][0]  # Single Meetings row
            ws.cell(row=task_row, column=units_col, value=meetings_hours)
        else:
            # Fill task rows for standard phases
            for i, task_code in enumerate(TASK_ORDER_STANDARD):
                task_hours = hours_dict.get((task_code, phase_code), 0)
                task_row = phase_info["tasks"][i]
                ws.cell(row=task_row, column=units_col, value=task_hours)


def delete_rows_from_sheet(ws, rows_to_delete: list[int]) -> int:
    """
    Delete rows from worksheet (must be sorted descending).

    Returns the number of rows deleted.
    """
    for row_idx in rows_to_delete:
        ws.delete_rows(row_idx, 1)
    return len(rows_to_delete)


def rebuild_formulas(ws) -> None:
    """
    Rebuild all formulas after row deletions.

    Scans the worksheet to find actual row positions and writes fresh formulas.
    This is needed because openpyxl's ArrayFormula objects don't update
    references when rows are deleted.
    """
    # Phase headers to look for
    phase_headers = {
        "Pre-Design", "Schematic Design", "Design Development",
        "Construction Documents", "Construction Administration",
        "Meetings w/ Client or Contractor"
    }
    # Task descriptions to look for
    task_descriptions = {
        "Design Principal", "Project Management", "3D Model",
        "Design and Documentation", "Meetings"
    }

    # First pass: find all phase headers and task rows
    phase_rows = []  # [(row, phase_name), ...]
    task_rows_map = {}  # row -> True for task rows

    for row in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row, column=2).value
        if not cell_b:
            continue
        cell_b_str = str(cell_b).strip()

        if cell_b_str in phase_headers:
            phase_rows.append(row)
        elif cell_b_str in task_descriptions:
            task_rows_map[row] = True

    # Find overall subtotal and reimbursable rows
    overall_subtotal_row = None
    reimbursable_subtotal_row = None
    reimbursable_cost_rows = []
    total_amount_due_row = None

    in_reimbursable = False
    for row in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row, column=2).value
        cell_d = ws.cell(row=row, column=4).value
        if cell_b:
            cell_b_str = str(cell_b).strip()
            if cell_b_str == "Reimbursable":
                in_reimbursable = True
            elif cell_b_str == "Subtotal":
                if not in_reimbursable:
                    overall_subtotal_row = row
                else:
                    reimbursable_subtotal_row = row
            elif cell_b_str in ["CCI Engineering", "Phipps Printing", "In house plotting(s.f.)"]:
                reimbursable_cost_rows.append(row)
        # Total Amount Due is in column D, not B
        if cell_d and "Total Amount Due" in str(cell_d):
            total_amount_due_row = row

    # Build phase structure: for each phase, find its task rows and subtotal row
    phase_task_rows = {}  # phase_row -> [task_rows]
    phase_subtotal_rows = {}  # phase_row -> subtotal_row

    for i, phase_row in enumerate(phase_rows):
        # Find the end boundary for this phase
        if i + 1 < len(phase_rows):
            next_phase_row = phase_rows[i + 1]
        else:
            next_phase_row = overall_subtotal_row or ws.max_row

        # Find task rows in this phase (between phase header and next phase/subtotal)
        phase_tasks = []
        subtotal_row = None

        for row in range(phase_row + 1, next_phase_row):
            cell_b = ws.cell(row=row, column=2).value
            if row in task_rows_map:
                phase_tasks.append(row)
            elif not cell_b:
                # Empty B cell after tasks - this is the phase subtotal row
                if phase_tasks and subtotal_row is None:
                    subtotal_row = row

        if phase_tasks:
            phase_task_rows[phase_row] = phase_tasks
            if subtotal_row:
                phase_subtotal_rows[phase_row] = subtotal_row

    # Now write formulas

    # 1. Cost formulas for each task row: E = C * D
    all_task_rows = []
    for task_rows in phase_task_rows.values():
        all_task_rows.extend(task_rows)
        for task_row in task_rows:
            ws.cell(row=task_row, column=5, value=f"=C{task_row}*D{task_row}")

    # 2. Phase subtotal formulas: F = SUM(E:E)
    phase_subtotal_list = []
    for phase_row, task_rows in phase_task_rows.items():
        subtotal_row = phase_subtotal_rows.get(phase_row)
        if subtotal_row and task_rows:
            first_task = min(task_rows)
            last_task = max(task_rows)
            ws.cell(row=subtotal_row, column=6, value=f"=SUM(E{first_task}:E{last_task})")
            phase_subtotal_list.append(subtotal_row)

    # 3. Overall subtotal formula
    if overall_subtotal_row:
        # Sum of all task Units (column C)
        if all_task_rows:
            units_parts = [f"C{r}" for r in sorted(all_task_rows)]
            ws.cell(row=overall_subtotal_row, column=3, value=f"=SUM({','.join(units_parts)})")

        # Sum of all phase subtotals (column F)
        if phase_subtotal_list:
            total_parts = [f"F{r}" for r in sorted(phase_subtotal_list)]
            ws.cell(row=overall_subtotal_row, column=6, value=f"=SUM({','.join(total_parts)})")

    # 4. Reimbursable subtotal formula
    if reimbursable_subtotal_row and reimbursable_cost_rows:
        first_reimb = min(reimbursable_cost_rows)
        last_reimb = max(reimbursable_cost_rows)
        ws.cell(row=reimbursable_subtotal_row, column=6, value=f"=SUM(E{first_reimb}:E{last_reimb})")

    # 5. Total Amount Due formula
    if total_amount_due_row and overall_subtotal_row:
        if reimbursable_subtotal_row:
            ws.cell(row=total_amount_due_row, column=6, value=f"=F{overall_subtotal_row}+F{reimbursable_subtotal_row}")
        else:
            ws.cell(row=total_amount_due_row, column=6, value=f"=F{overall_subtotal_row}")


def create_timesheet_sheet(wb, sheet_name: str, entries: list[dict]) -> None:
    """
    Create a separate timesheet worksheet with entries and total row.
    """
    ws = wb.create_sheet(title=sheet_name)

    # Turn off gridlines
    ws.sheet_view.showGridLines = False

    # Write header with styling (bold white text + gray background)
    # Column order: Date, E (Employee), H (Hours), P (Phase), T (Task), WID
    headers = ["Date", "E", "H", "P", "T", "WID"]
    header_font = Font(bold=True, color=Color(rgb='FFFFFFFF'))
    header_fill = PatternFill(patternType='solid', fgColor=Color(indexed=11))
    header_row = 1

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data rows
    first_data_row = header_row + 1
    for row_offset, entry in enumerate(entries):
        data_row = first_data_row + row_offset

        date_str = ""
        if entry["date"]:
            date_str = format_timesheet_date(entry["date"])

        ws.cell(row=data_row, column=1, value=date_str)
        ws.cell(row=data_row, column=2, value=entry["employee"])
        ws.cell(row=data_row, column=3, value=entry["hours"])
        ws.cell(row=data_row, column=4, value=entry["phase"])
        ws.cell(row=data_row, column=5, value=entry["task"])
        ws.cell(row=data_row, column=6, value=entry["wid"])

    # Add total row
    last_data_row = first_data_row + len(entries) - 1
    total_row = last_data_row + 1

    total_cell = ws.cell(row=total_row, column=1, value="Total")
    total_cell.font = Font(bold=True)

    hours_sum_cell = ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})")
    hours_sum_cell.font = Font(bold=True)

    # Set optimized column widths for portrait page fit
    column_widths = {
        'A': 30.6640625, # Date
        'B': 4.33203125, # E (Employee)
        'C': 4.1640625,  # H (Hours)
        'D': 3.5,        # P (Phase)
        'E': 4.0,        # T (Task)
        'F': 37.0,       # WID
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width


def apply_style_fixes(ws) -> None:
    """
    Apply style fixes to the worksheet.

    - Insert blank row under logo
    - Turn off gridlines
    - Add logo to top-left corner
    - Set row height for footer row (to accommodate multi-line text)
    """
    # Insert blank row under logo (at row 2)
    ws.insert_rows(2, 1)

    # Turn off gridlines
    ws.sheet_view.showGridLines = False

    # Add logo to top-left corner (A1)
    if LOGO_PATH.exists():
        logo = ExcelImage(str(LOGO_PATH))
        # Scale logo to reasonable size (adjust as needed)
        logo.width = 150
        logo.height = 40
        ws.add_image(logo, "A1")

    # Find footer row (contains "CCI Design Inc.") and set row height
    for row in range(1, ws.max_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and "CCI Design Inc." in cell.value:
                # Set row height to accommodate 2 lines of text
                ws.row_dimensions[row].height = 30
                return


def process_project_sheet(
    ws,
    project_id: str,
    hours_dict: dict[tuple[str, str], float],
) -> None:
    """Process a single project invoice worksheet."""
    project_name, project_number = parse_project_id(project_id)

    # Calculate rows to delete
    rows_to_delete = calculate_rows_to_delete(hours_dict)

    # Populate invoice data (before deleting rows)
    populate_invoice_sheet(ws, project_name, project_number, hours_dict)

    # Delete empty rows (from bottom to top)
    delete_rows_from_sheet(ws, rows_to_delete)

    # Apply style fixes (inserts blank row under logo, gridlines off, logo, footer row height)
    apply_style_fixes(ws)

    # Rebuild formulas with correct cell references after row deletions and blank row insertion
    rebuild_formulas(ws)


def create_invoice_workbook(
    template_path: Path,
    aggregated_data: dict[str, dict[tuple[str, str], float]],
    detail_entries: dict[str, list[dict]],
) -> any:
    """Create the output Excel workbook with all project sheets."""
    print(f"Loading template: {template_path}")

    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    # Load template
    wb = load_workbook(str(template_path))
    template_ws = wb.active
    template_name = template_ws.title

    # Sort projects alphabetically by name (part before colon)
    sorted_projects = sorted(
        aggregated_data.keys(),
        key=lambda p: p.split(":")[0].strip().lower(),
    )

    # Process each project - create invoice sheet and timesheet sheet
    for project_id in sorted_projects:
        hours = dict(aggregated_data[project_id])
        entries = detail_entries[project_id]

        phase_totals = calculate_phase_hours(hours)
        total_hours = sum(phase_totals.values())
        active_phases = [p for p in PHASE_ORDER if phase_totals.get(p, 0) > 0]
        removed_phases = [p for p in PHASE_ORDER if phase_totals.get(p, 0) == 0]

        print(f"\nProcessing: {project_id}")
        print(f"  - {len(entries)} entries, {total_hours:.1f} total hours")
        if active_phases:
            phase_str = ", ".join(f"{p} ({phase_totals[p]:.1f}h)" for p in active_phases)
            print(f"  - Phases: {phase_str}")
        if removed_phases:
            print(f"  - Removing empty phases: {', '.join(removed_phases)}")

        # Create invoice sheet (copy from template)
        invoice_ws = wb.copy_worksheet(template_ws)
        invoice_ws.title = make_sheet_name(project_id, " A")
        process_project_sheet(invoice_ws, project_id, hours)

        # Create timesheet sheet
        timesheet_name = make_sheet_name(project_id, " B")
        create_timesheet_sheet(wb, timesheet_name, entries)

    # Remove the original template sheet (it was only used as a source)
    del wb[template_name]

    return wb


# =============================================================================
# MAIN
# =============================================================================


def main(input_file: Path) -> None:
    """Main entry point for invoice generation."""
    # Read input data
    all_entries = read_timesheet_data(input_file)
    unique_projects = set(e["project_id"] for e in all_entries)
    print(f"Found {len(all_entries)} entries across {len(unique_projects)} unique Project IDs")

    # Filter non-projects
    filtered_entries, excluded_count = filter_non_projects(all_entries)
    if excluded_count > 0:
        print(f"Filtering non-projects... {excluded_count} excluded (Office, Vacation, Holiday, etc.)")

    # Filter zero hours
    filtered_entries = filter_zero_hours(filtered_entries)

    # Validate project IDs
    project_errors = validate_project_ids(filtered_entries)
    if project_errors:
        print("\nERROR: Invalid Project IDs found:")
        for error in project_errors:
            print(f"  - {error}")
        sys.exit(1)

    # Validate codes
    code_errors = validate_codes(filtered_entries)
    if code_errors:
        print("\nERROR: Invalid Task/Phase codes found:")
        for error in code_errors:
            print(f"  - {error}")
        sys.exit(1)

    print("Validating codes... OK")

    # Check for billable projects
    if not filtered_entries:
        print("\nERROR: No billable projects found in input file")
        sys.exit(1)

    # Aggregate and group
    aggregated = aggregate_hours(filtered_entries)
    grouped = group_by_project(filtered_entries)

    print(f"{len(aggregated)} billable projects to process")

    # Create invoice workbook
    wb = create_invoice_workbook(TEMPLATE_PATH, aggregated, grouped)

    # Generate output filename and save
    output_file = generate_output_filename(input_file, filtered_entries)
    print(f"\nSorting sheets alphabetically...")
    print(f"Writing output: {output_file}")

    wb.save(str(output_file))

    # Summary
    total_hours = sum(
        sum(hours.values()) for hours in aggregated.values()
    )
    print(f"\nComplete! Generated invoices for {len(aggregated)} projects ({total_hours:.1f} total hours)")
    print(f"Output: {output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate client invoices from monthly timesheet reports"
    )
    parser.add_argument(
        "input_file",
        type=Path,
        help="Path to the monthly report Numbers file",
    )

    args = parser.parse_args()
    main(args.input_file)
