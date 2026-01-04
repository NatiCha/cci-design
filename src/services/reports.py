"""
Report generation utilities for Numbers and Excel formats.
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from core.config import (
    BILLABLE_GOALS_ROW_LABELS,
    DETAIL_HEADERS,
    DETAIL_HEADERS_EDIT,
    NON_PROJECT_NAMES,
)


def format_date_display(d: date) -> str:
    """Format date as M/D/YYYY (platform-safe, no zero-padding)."""
    return f"{d.month}/{d.day}/{d.year}"


def format_date_short(d: date) -> str:
    """Format date as 'Mon D' (platform-safe, e.g., 'Nov 7')."""
    return f"{d.strftime('%b')} {d.day}"


def format_date_for_subject(d: date, report_type: str) -> str:
    """Format date for email subject."""
    if report_type == "weekly_report":
        # "Nov 7th 2025"
        day = d.day
        suffix = "th" if 11 <= day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        return d.strftime(f"%b {day}{suffix} %Y")
    else:
        # "November 2025"
        return d.strftime("%B %Y")


def write_detail_table(table, events: list[dict]):
    """
    Write headers and event data to a Numbers detail table.

    Reusable helper for both weekly and monthly reports.
    """
    # Write headers
    for col_idx, header in enumerate(DETAIL_HEADERS):
        table.write(0, col_idx, header)

    # Write data rows
    for row_idx, event in enumerate(events, start=1):
        date_str = ""
        if event["event_date"]:
            date_str = format_date_display(event["event_date"])

        row_data = [
            event["project_id"],
            date_str,
            event["employee_id"].upper(),
            event["hours"],
            event["task"].upper() if event["task"] else "",
            event["phase"].upper() if event["phase"] else "",
            event["wid"] or "",
        ]

        for col_idx, value in enumerate(row_data):
            table.write(row_idx, col_idx, value)


# =============================================================================
# EXCEL REPORT GENERATION (Monthly Reports)
# =============================================================================


def write_excel_detail_view_sheet(ws, events: list[dict]):
    """
    Write Sheet 1 - Timesheet Detail (View) to Excel worksheet.

    Uses standard DETAIL_HEADERS: Project ID, Date, Employee, Hours, Task, Phase, WID
    """
    # Write headers (row 1)
    for col_idx, header in enumerate(DETAIL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Write data rows
    for row_idx, event in enumerate(events, start=2):
        date_str = format_date_display(event["event_date"]) if event["event_date"] else ""

        row_data = [
            event["project_id"],
            date_str,
            event["employee_id"].upper(),
            event["hours"],
            event["task"].upper() if event["task"] else "",
            event["phase"].upper() if event["phase"] else "",
            event["wid"] or "",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def write_excel_detail_edit_sheet(ws, events: list[dict]):
    """
    Write Sheet 2 - Timesheet Detail (Edit) to Excel worksheet.

    Headers: Project ID, Date, Employee, Hours, Hours Adjusted, Total Adjusted Hours, Task, Phase, WID
    Column E (Hours Adjusted): Empty for user input
    Column F (Total Adjusted Hours): Formula =D{row}+IF(E{row}="",0,E{row})
    """
    # Write headers (row 1)
    for col_idx, header in enumerate(DETAIL_HEADERS_EDIT, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Write data rows
    for row_idx, event in enumerate(events, start=2):
        date_str = format_date_display(event["event_date"]) if event["event_date"] else ""

        # Columns A-D: Project ID, Date, Employee, Hours
        ws.cell(row=row_idx, column=1, value=event["project_id"])
        ws.cell(row=row_idx, column=2, value=date_str)
        ws.cell(row=row_idx, column=3, value=event["employee_id"].upper())
        ws.cell(row=row_idx, column=4, value=event["hours"])

        # Column E (Hours Adjusted): Leave empty for user input

        # Column F (Total Adjusted Hours): Formula
        ws.cell(row=row_idx, column=6, value=f'=D{row_idx}+IF(E{row_idx}="",0,E{row_idx})')

        # Columns G-I: Task, Phase, WID
        ws.cell(row=row_idx, column=7, value=event["task"].upper() if event["task"] else "")
        ws.cell(row=row_idx, column=8, value=event["phase"].upper() if event["phase"] else "")
        ws.cell(row=row_idx, column=9, value=event["wid"] or "")


def write_excel_billable_goals_sheet(ws, events: list[dict], edit_sheet_name: str):
    """
    Write Sheet 3 - Billable Goals to Excel worksheet.

    Dynamically determines employee columns from events.
    Uses SUMIF formulas referencing the Edit sheet.

    Structure:
    Row 1: Headers - (empty) | Employee1 | Employee2 | ... | Total
    Row 2: Billable goal (empty cells for user input, Total = SUM)
    Row 3: Gross billable hours adjusted (SUMIF formulas)
    Row 4: % of goal (= Row3 / Row2)
    Row 5: Non-project hours (SUMIFS for non-projects)
    Row 6: Net billable hours this period (= Row3 - Row5)
    """
    # Get unique employees sorted
    employees = sorted(set(e["employee_id"].upper() for e in events))

    # Calculate data range in Edit sheet
    data_start_row = 2
    data_end_row = len(events) + 1

    # Column references for Edit sheet (letters)
    # A=Project ID, B=Date, C=Employee, D=Hours, E=Hours Adjusted, F=Total Adjusted Hours
    employee_col = "C"
    total_adj_hours_col = "F"
    project_id_col = "A"

    # Write headers
    headers = [""] + employees + ["Total"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Column letters for employee columns (B, C, D, etc.)
    first_emp_col = get_column_letter(2)
    last_emp_col = get_column_letter(len(employees) + 1)
    total_col = len(employees) + 2

    # Row 2: Billable goal
    ws.cell(row=2, column=1, value=BILLABLE_GOALS_ROW_LABELS[0])
    # Employee columns: empty for user input
    # Total column: SUM formula
    ws.cell(row=2, column=total_col, value=f"=SUM({first_emp_col}2:{last_emp_col}2)")

    # Row 3: Gross billable hours adjusted
    ws.cell(row=3, column=1, value=BILLABLE_GOALS_ROW_LABELS[1])
    for emp_idx, emp in enumerate(employees):
        col = emp_idx + 2
        # SUMIF: sum Total Adjusted Hours where Employee matches
        formula = (
            f"=SUMIF('{edit_sheet_name}'!${employee_col}${data_start_row}:${employee_col}${data_end_row},"
            f'"{emp}",'
            f"'{edit_sheet_name}'!${total_adj_hours_col}${data_start_row}:${total_adj_hours_col}${data_end_row})"
        )
        ws.cell(row=3, column=col, value=formula)
    # Total column: SUM
    ws.cell(row=3, column=total_col, value=f"=SUM({first_emp_col}3:{last_emp_col}3)")

    # Row 4: % of goal
    ws.cell(row=4, column=1, value=BILLABLE_GOALS_ROW_LABELS[2])
    for emp_idx, emp in enumerate(employees):
        col = emp_idx + 2
        col_letter = get_column_letter(col)
        # Avoid division by zero with IFERROR
        ws.cell(row=4, column=col, value=f"=IFERROR({col_letter}3/{col_letter}2,0)")
    # Total column: AVERAGE
    ws.cell(row=4, column=total_col, value=f"=AVERAGE({first_emp_col}4:{last_emp_col}4)")

    # Row 5: Non-project hours
    # Use SUMIFS for each non-project type with wildcard matching
    ws.cell(row=5, column=1, value=BILLABLE_GOALS_ROW_LABELS[3])
    for emp_idx, emp in enumerate(employees):
        col = emp_idx + 2
        # Build sum of SUMIFS for each non-project type
        non_project_parts = []
        for np_name in sorted(NON_PROJECT_NAMES):
            # SUMIFS for each non-project type (case-insensitive matching with wildcard)
            sumif = (
                f"SUMIFS('{edit_sheet_name}'!${total_adj_hours_col}${data_start_row}:${total_adj_hours_col}${data_end_row},"
                f"'{edit_sheet_name}'!${employee_col}${data_start_row}:${employee_col}${data_end_row},"
                f'"{emp}",'
                f"'{edit_sheet_name}'!${project_id_col}${data_start_row}:${project_id_col}${data_end_row},"
                f'"{np_name}:*")'
            )
            non_project_parts.append(sumif)

        formula = "=" + "+".join(non_project_parts)
        ws.cell(row=5, column=col, value=formula)
    # Total column: SUM
    ws.cell(row=5, column=total_col, value=f"=SUM({first_emp_col}5:{last_emp_col}5)")

    # Row 6: Net billable hours this period (Gross - Non-project)
    ws.cell(row=6, column=1, value=BILLABLE_GOALS_ROW_LABELS[4])
    for emp_idx, emp in enumerate(employees):
        col = emp_idx + 2
        col_letter = get_column_letter(col)
        ws.cell(row=6, column=col, value=f"={col_letter}3-{col_letter}5")
    # Total column: SUM
    ws.cell(row=6, column=total_col, value=f"=SUM({first_emp_col}6:{last_emp_col}6)")


def create_monthly_excel_report(events: list[dict], output_path: Path):
    """
    Create Excel monthly report with three sheets.

    Sheet 1: "1 Timesheet Detail (View)" - Standard 7 columns
    Sheet 2: "3 Timesheet Detail (Edit)" - 9 columns with Hours Adjusted formulas
    Sheet 3: "4 Billable Goals" - Dynamic employee columns with SUMIF formulas
    """
    wb = Workbook()

    # Sheet 1: Timesheet Detail (View)
    ws_view = wb.active
    ws_view.title = "1 Timesheet Detail (View)"
    write_excel_detail_view_sheet(ws_view, events)

    # Sheet 2: Timesheet Detail (Edit)
    ws_edit = wb.create_sheet(title="3 Timesheet Detail (Edit)")
    write_excel_detail_edit_sheet(ws_edit, events)

    # Sheet 3: Billable Goals
    ws_goals = wb.create_sheet(title="4 Billable Goals")
    write_excel_billable_goals_sheet(ws_goals, events, "3 Timesheet Detail (Edit)")

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Saved Excel report to: {output_path}")
