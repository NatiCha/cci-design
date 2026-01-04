"""
Invoice Generation Service (Excel Version)

Generates client invoices from monthly timesheet reports. Reads timesheet data
from an Apple Numbers file, uses an Excel template to create professionally formatted
invoice cover sheets per project with Excel formulas, and attaches detailed timesheet breakdowns.
"""

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from numbers_parser import Document
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Color, Font, PatternFill

from core.config import NON_PROJECT_NAMES, OUTPUT_DIR, TEMPLATES_DIR


# =============================================================================
# DATA CLASSES
# =============================================================================


@dataclass
class InvoiceResult:
    """Result of invoice generation processing."""

    workbook: Any  # openpyxl.Workbook
    project_count: int
    total_hours: float
    entries: list[dict]

# =============================================================================
# CONSTANTS
# =============================================================================

TEMPLATE_PATH = TEMPLATES_DIR / "invoice-template.xlsx"
LOGO_PATH = TEMPLATES_DIR / "cci-design-logo.png"

# Phase code to template description mapping
PHASE_TO_DESCRIPTION = {
    "PD": "Pre-Design",
    "SD": "Schematic Design",
    "DD": "Design Development",
    "CD": "Construction Documents",
    "CA": "Construction Administration",
    "M": "Meetings w/ Client or Contractor",
}

# Task code to template description mapping
TASK_TO_DESCRIPTION = {
    "DP": "Design Principal",
    "PM": "Project Management",
    "3-D": "3D Model",
    "D-D": "Design and Documentation",
    "M": "Meetings",
}

# Processing order
PHASE_ORDER = ["PD", "SD", "DD", "CD", "CA", "M"]
TASK_ORDER_STANDARD = ["DP", "PM", "3-D", "D-D"]
TASK_ORDER_MEETINGS = ["M"]

# Valid codes for billable entries (excludes BD, NA which are for non-projects)
BILLABLE_TASK_CODES = {"DP", "PM", "3-D", "D-D", "M"}
BILLABLE_PHASE_CODES = {"PD", "SD", "DD", "CD", "CA", "M"}

# Template row indices (1-indexed for Excel)
TEMPLATE_STRUCTURE = {
    "project_name_row": 3,
    "project_number_row": 4,
    "invoice_date_row": 5,
    "invoice_number_row": 6,
    "value_col": 3,  # Column C for placeholder values
    "units_col": 3,  # Column C for Units
    "rate_col": 4,   # Column D for Rate
    "cost_col": 5,   # Column E for Cost
    "total_col": 6,  # Column F for Total
    "phases": {
        "PD": {"header": 10, "tasks": [11, 12, 13, 14], "subtotal": 15},
        "SD": {"header": 16, "tasks": [17, 18, 19, 20], "subtotal": 21},
        "DD": {"header": 22, "tasks": [23, 24, 25, 26], "subtotal": 27},
        "CD": {"header": 28, "tasks": [29, 30, 31, 32], "subtotal": 33},
        "CA": {"header": 34, "tasks": [35, 36, 37, 38], "subtotal": 39},
        "M": {"header": 40, "tasks": [41], "subtotal": 42},
    },
    "overall_subtotal_row": 43,
    "reimbursable_header": 44,
    "reimbursable_rows": [45, 46, 47],
    "reimbursable_subtotal": 48,
    "total_amount_due_row": 50,
    "footer_row": 52,
}

# Task code to row offset within phase section (for standard phases)
TASK_TO_ROW_OFFSET = {"DP": 0, "PM": 1, "3-D": 2, "D-D": 3}

# Excel sheet name limit
MAX_SHEET_NAME_LENGTH = 31


# =============================================================================
# INPUT READING
# =============================================================================


def read_timesheet_data(input_file: Path) -> list[dict]:
    """
    Read timesheet entries from the input Numbers file.

    Reads from sheet "3 Timesheet Detail (Edit)", table "Timesheet Detail (Edit)".

    The Edit sheet has 9 columns:
    - A: Project ID
    - B: Date
    - C: Employee
    - D: Hours (original)
    - E: Hours Adjusted (user input, may be empty)
    - F: Total Adjusted Hours (formula: Hours + Hours Adjusted)
    - G: Task
    - H: Phase
    - I: WID

    For invoices, we use "Total Adjusted Hours" (column F) as the hours value.

    Returns:
        List of entry dicts with keys: project_id, date, employee, hours, task, phase, wid
    """
    print(f"Reading input file: {input_file}")

    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    doc = Document(str(input_file))

    sheet_name = "3 Timesheet Detail (Edit)"
    table_name = "Table 1"

    try:
        sheet = doc.sheets[sheet_name]
        table = sheet.tables[table_name]
    except (KeyError, IndexError) as e:
        raise ValueError(
            f"Expected sheet '{sheet_name}' with table '{table_name}' not found in input file"
        ) from e

    print(f"Parsing timesheet data from '{sheet_name}'...")

    rows = table.rows(values_only=True)
    if len(rows) < 2:
        raise ValueError("Input file contains no data rows")

    # New column structure (9 columns):
    # 0: Project ID, 1: Date, 2: Employee, 3: Hours, 4: Hours Adjusted,
    # 5: Total Adjusted Hours, 6: Task, 7: Phase, 8: WID
    entries = []
    for row in rows[1:]:  # Skip header
        if len(row) < 9:
            continue

        project_id = str(row[0]).strip() if row[0] else ""
        date_val = row[1]
        employee = str(row[2]).strip() if row[2] else ""
        # Use Total Adjusted Hours (column F, index 5) for invoice hours
        # This reflects any manual adjustments the user made
        total_adjusted_hours = float(row[5]) if row[5] else 0.0
        task = str(row[6]).strip().upper() if row[6] else ""
        phase = str(row[7]).strip().upper() if row[7] else ""
        wid = str(row[8]).strip() if row[8] else ""

        # Convert date to date object if needed
        if isinstance(date_val, datetime):
            date_val = date_val.date()
        elif isinstance(date_val, str):
            try:
                date_val = datetime.strptime(date_val, "%m/%d/%Y").date()
            except ValueError:
                date_val = None

        entries.append(
            {
                "project_id": project_id,
                "date": date_val,
                "employee": employee,
                "hours": total_adjusted_hours,
                "task": task,
                "phase": phase,
                "wid": wid,
            }
        )

    return entries


def parse_project_id(project_id: str) -> tuple[str, str]:
    """
    Parse project ID into name and number.

    Args:
        project_id: String in format "Name: Number"

    Returns:
        Tuple of (project_name, project_number)

    Raises:
        ValueError: If project_id doesn't contain colon separator
    """
    if ":" not in project_id:
        raise ValueError(f"Project ID missing colon separator: '{project_id}'")

    parts = project_id.split(":", 1)
    name = parts[0].strip()
    number = parts[1].strip()
    return name, number


# =============================================================================
# FILTERING & VALIDATION
# =============================================================================


def filter_non_projects(entries: list[dict]) -> tuple[list[dict], int]:
    """
    Remove non-project entries (Office, Vacation, Holiday, Sick, Personal Time).

    Returns:
        Tuple of (filtered entries, count of excluded entries)
    """
    filtered = []
    excluded_count = 0

    for entry in entries:
        project_id = entry["project_id"].lower()
        is_non_project = False

        for non_project in NON_PROJECT_NAMES:
            if project_id.startswith(non_project + ":") or project_id == non_project:
                is_non_project = True
                break

        if is_non_project:
            excluded_count += 1
        else:
            filtered.append(entry)

    return filtered, excluded_count


def filter_zero_hours(entries: list[dict]) -> list[dict]:
    """Remove entries with zero or negative hours."""
    return [e for e in entries if e["hours"] > 0]


def validate_codes(entries: list[dict]) -> list[str]:
    """
    Validate Task and Phase codes against allowed billable codes.

    Returns:
        List of error messages (empty if all valid)
    """
    errors = []

    for entry in entries:
        task = entry["task"]
        phase = entry["phase"]
        project_id = entry["project_id"]

        if task and task not in BILLABLE_TASK_CODES:
            errors.append(
                f"Invalid Task code '{task}' for project '{project_id}' "
                f"(valid: {', '.join(sorted(BILLABLE_TASK_CODES))})"
            )

        if phase and phase not in BILLABLE_PHASE_CODES:
            errors.append(
                f"Invalid Phase code '{phase}' for project '{project_id}' "
                f"(valid: {', '.join(sorted(BILLABLE_PHASE_CODES))})"
            )

    return errors


def validate_project_ids(entries: list[dict]) -> list[str]:
    """
    Validate all Project IDs can be parsed (contain colon).

    Returns:
        List of error messages for unparseable project IDs
    """
    errors = []
    seen = set()

    for entry in entries:
        project_id = entry["project_id"]
        if project_id in seen:
            continue
        seen.add(project_id)

        if ":" not in project_id:
            errors.append(f"Project ID missing colon separator: '{project_id}'")

    return errors


# =============================================================================
# AGGREGATION
# =============================================================================


def aggregate_hours(entries: list[dict]) -> dict[str, dict[tuple[str, str], float]]:
    """
    Aggregate hours by project, then by (task, phase) combination.

    Returns:
        Nested dict: {project_id: {(task_code, phase_code): total_hours}}
    """
    result = defaultdict(lambda: defaultdict(float))

    for entry in entries:
        project_id = entry["project_id"]
        task = entry["task"]
        phase = entry["phase"]
        hours = entry["hours"]

        if task and phase:
            result[project_id][(task, phase)] += hours

    return result


def group_by_project(entries: list[dict]) -> dict[str, list[dict]]:
    """
    Group individual entries by project for detail tables.

    Returns:
        Dict mapping project_id to list of entries, sorted by date
    """
    grouped = defaultdict(list)

    for entry in entries:
        grouped[entry["project_id"]].append(entry)

    # Sort each project's entries by date
    for project_id in grouped:
        grouped[project_id].sort(key=lambda e: e["date"] or date.min)

    return grouped


# =============================================================================
# OUTPUT GENERATION
# =============================================================================


def generate_output_filename(input_file: Path, entries: list[dict]) -> Path:
    """
    Generate output filename with versioning.

    Format: invoices_YYYY_MM_a.xlsx (e.g., invoices_2025_11_a.xlsx)
    Adds _a, _b, _c suffix if file exists for proper alphabetical sorting.
    """
    # Extract year/month from entries
    for entry in entries:
        if entry["date"]:
            year = entry["date"].year
            month = entry["date"].month
            break
    else:
        # Fallback to current date
        today = date.today()
        year = today.year
        month = today.month

    output_dir = OUTPUT_DIR / "invoices"
    output_dir.mkdir(parents=True, exist_ok=True)
    base_name = f"invoices_{year}_{month:02d}"

    # Check for existing files and add suffix
    suffix_char = ord("a")

    while True:
        suffix = f"_{chr(suffix_char)}"
        filename = f"{base_name}{suffix}.xlsx"
        output_path = output_dir / filename
        if not output_path.exists():
            return output_path
        suffix_char += 1
        if suffix_char > ord("z"):
            raise RuntimeError("Too many output files exist")


def ordinal_suffix(day: int) -> str:
    """Return ordinal suffix for a day number (1st, 2nd, 3rd, etc.)."""
    if 11 <= day <= 13:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")


def get_invoice_date(override: date | None = None) -> str:
    """Get invoice date formatted as 'December 30, 2025'.

    Args:
        override: Optional date to use instead of current date

    Returns:
        Formatted date string
    """
    if override:
        return override.strftime("%B %d, %Y").replace(" 0", " ")
    tz = ZoneInfo("America/New_York")
    today = datetime.now(tz).date()
    return today.strftime("%B %d, %Y").replace(" 0", " ")


def format_timesheet_date(d: date) -> str:
    """Format date as 'Thursday, October 30th 2025'."""
    day_suffix = ordinal_suffix(d.day)
    base = d.strftime(f"%A, %B {d.day}{day_suffix} %Y")
    return base


def make_sheet_name(name: str, suffix: str) -> str:
    """Create a valid Excel sheet name with suffix (31 chars max)."""
    # Replace colon with space-hyphen-space for readability
    sanitized = name.replace(":", " -")

    # Replace other invalid characters with hyphen
    invalid_chars = ["\\", "/", "?", "*", "[", "]"]
    for char in invalid_chars:
        sanitized = sanitized.replace(char, "-")

    # Calculate max length for the base name
    max_base_len = MAX_SHEET_NAME_LENGTH - len(suffix)

    # Truncate base name if needed
    if len(sanitized) > max_base_len:
        sanitized = sanitized[:max_base_len].rstrip()

    return sanitized + suffix


# =============================================================================
# EXCEL MANIPULATION
# =============================================================================


def calculate_phase_hours(hours_dict: dict[tuple[str, str], float]) -> dict[str, float]:
    """Calculate total hours per phase."""
    phase_totals = defaultdict(float)
    for (_, phase), hours in hours_dict.items():
        phase_totals[phase] += hours
    return phase_totals


def calculate_rows_to_delete(hours_dict: dict[tuple[str, str], float]) -> list[int]:
    """
    Calculate which rows to delete, returned in descending order.

    Deletes:
    - Entire phase sections if phase has 0 hours
    - Individual task rows with 0 hours in phases that have some hours
    """
    rows_to_delete = []
    phase_totals = calculate_phase_hours(hours_dict)

    for phase_code in PHASE_ORDER:
        phase_info = TEMPLATE_STRUCTURE["phases"][phase_code]
        phase_hours = phase_totals.get(phase_code, 0)

        if phase_hours == 0:
            # Delete entire phase section
            rows_to_delete.append(phase_info["header"])
            rows_to_delete.extend(phase_info["tasks"])
            rows_to_delete.append(phase_info["subtotal"])
        elif phase_code != "M":
            # For non-Meetings phases, delete individual zero-hour task rows
            for i, task_code in enumerate(TASK_ORDER_STANDARD):
                task_hours = hours_dict.get((task_code, phase_code), 0)
                if task_hours == 0:
                    rows_to_delete.append(phase_info["tasks"][i])

    return sorted(rows_to_delete, reverse=True)


def populate_invoice_sheet(
    ws,
    project_name: str,
    project_number: str,
    hours_dict: dict[tuple[str, str], float],
    invoice_date: date | None = None,
) -> None:
    """Populate the invoice worksheet with project data."""
    value_col = TEMPLATE_STRUCTURE["value_col"]
    units_col = TEMPLATE_STRUCTURE["units_col"]

    # Fill header fields
    ws.cell(row=TEMPLATE_STRUCTURE["project_name_row"], column=value_col, value=project_name)
    ws.cell(row=TEMPLATE_STRUCTURE["project_number_row"], column=value_col, value=project_number)
    ws.cell(row=TEMPLATE_STRUCTURE["invoice_date_row"], column=value_col, value=get_invoice_date(invoice_date))
    ws.cell(row=TEMPLATE_STRUCTURE["invoice_number_row"], column=value_col, value=f"{project_number} INV ##")

    # Fill hours for each phase/task
    for phase_code in PHASE_ORDER:
        phase_info = TEMPLATE_STRUCTURE["phases"][phase_code]

        if phase_code == "M":
            # For Meetings phase, aggregate ALL tasks into the single Meetings row
            meetings_hours = sum(
                hours for (task, phase), hours in hours_dict.items() if phase == "M"
            )
            task_row = phase_info["tasks"][0]
            ws.cell(row=task_row, column=units_col, value=meetings_hours)
        else:
            # Fill task rows for standard phases
            for i, task_code in enumerate(TASK_ORDER_STANDARD):
                task_hours = hours_dict.get((task_code, phase_code), 0)
                task_row = phase_info["tasks"][i]
                ws.cell(row=task_row, column=units_col, value=task_hours)


def delete_rows_from_sheet(ws, rows_to_delete: list[int]) -> int:
    """Delete rows from worksheet (must be sorted descending)."""
    for row_idx in rows_to_delete:
        ws.delete_rows(row_idx, 1)
    return len(rows_to_delete)


def rebuild_formulas(ws) -> None:
    """
    Rebuild all formulas after row deletions.

    Scans the worksheet to find actual row positions and writes fresh formulas.
    """
    # Phase headers to look for
    phase_headers = {
        "Pre-Design",
        "Schematic Design",
        "Design Development",
        "Construction Documents",
        "Construction Administration",
        "Meetings w/ Client or Contractor",
    }
    # Task descriptions to look for
    task_descriptions = {
        "Design Principal",
        "Project Management",
        "3D Model",
        "Design and Documentation",
        "Meetings",
    }

    # First pass: find all phase headers and task rows
    phase_rows = []
    task_rows_map = {}

    for row in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row, column=2).value
        if not cell_b:
            continue
        cell_b_str = str(cell_b).strip()

        if cell_b_str in phase_headers:
            phase_rows.append(row)
        elif cell_b_str in task_descriptions:
            task_rows_map[row] = True

    # Find overall subtotal and reimbursable rows
    overall_subtotal_row = None
    reimbursable_subtotal_row = None
    reimbursable_cost_rows = []
    total_amount_due_row = None

    in_reimbursable = False
    for row in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row, column=2).value
        cell_d = ws.cell(row=row, column=4).value
        if cell_b:
            cell_b_str = str(cell_b).strip()
            if cell_b_str == "Reimbursable":
                in_reimbursable = True
            elif cell_b_str == "Subtotal":
                if not in_reimbursable:
                    overall_subtotal_row = row
                else:
                    reimbursable_subtotal_row = row
            elif cell_b_str in ["CCI Engineering", "Phipps Printing", "In house plotting(s.f.)"]:
                reimbursable_cost_rows.append(row)
        if cell_d and "Total Amount Due" in str(cell_d):
            total_amount_due_row = row

    # Build phase structure
    phase_task_rows = {}
    phase_subtotal_rows = {}

    for i, phase_row in enumerate(phase_rows):
        if i + 1 < len(phase_rows):
            next_phase_row = phase_rows[i + 1]
        else:
            next_phase_row = overall_subtotal_row or ws.max_row

        phase_tasks = []
        subtotal_row = None

        for row in range(phase_row + 1, next_phase_row):
            cell_b = ws.cell(row=row, column=2).value
            if row in task_rows_map:
                phase_tasks.append(row)
            elif not cell_b:
                if phase_tasks and subtotal_row is None:
                    subtotal_row = row

        if phase_tasks:
            phase_task_rows[phase_row] = phase_tasks
            if subtotal_row:
                phase_subtotal_rows[phase_row] = subtotal_row

    # Write formulas

    # 1. Cost formulas for each task row: E = C * D
    all_task_rows = []
    for task_rows in phase_task_rows.values():
        all_task_rows.extend(task_rows)
        for task_row in task_rows:
            ws.cell(row=task_row, column=5, value=f"=C{task_row}*D{task_row}")

    # 2. Phase subtotal formulas: F = SUM(E:E)
    phase_subtotal_list = []
    for phase_row, task_rows in phase_task_rows.items():
        subtotal_row = phase_subtotal_rows.get(phase_row)
        if subtotal_row and task_rows:
            first_task = min(task_rows)
            last_task = max(task_rows)
            ws.cell(row=subtotal_row, column=6, value=f"=SUM(E{first_task}:E{last_task})")
            phase_subtotal_list.append(subtotal_row)

    # 3. Overall subtotal formula
    if overall_subtotal_row:
        if all_task_rows:
            units_parts = [f"C{r}" for r in sorted(all_task_rows)]
            ws.cell(row=overall_subtotal_row, column=3, value=f"=SUM({','.join(units_parts)})")

        if phase_subtotal_list:
            total_parts = [f"F{r}" for r in sorted(phase_subtotal_list)]
            ws.cell(row=overall_subtotal_row, column=6, value=f"=SUM({','.join(total_parts)})")

    # 4. Reimbursable subtotal formula
    if reimbursable_subtotal_row and reimbursable_cost_rows:
        first_reimb = min(reimbursable_cost_rows)
        last_reimb = max(reimbursable_cost_rows)
        ws.cell(row=reimbursable_subtotal_row, column=6, value=f"=SUM(E{first_reimb}:E{last_reimb})")

    # 5. Total Amount Due formula
    if total_amount_due_row and overall_subtotal_row:
        if reimbursable_subtotal_row:
            ws.cell(
                row=total_amount_due_row,
                column=6,
                value=f"=F{overall_subtotal_row}+F{reimbursable_subtotal_row}",
            )
        else:
            ws.cell(row=total_amount_due_row, column=6, value=f"=F{overall_subtotal_row}")


def create_timesheet_sheet(wb, sheet_name: str, entries: list[dict]) -> None:
    """Create a separate timesheet worksheet with entries and total row."""
    ws = wb.create_sheet(title=sheet_name)

    # Turn off gridlines
    ws.sheet_view.showGridLines = False

    # Write header with styling
    headers = ["Date", "E", "H", "P", "T", "WID"]
    header_font = Font(bold=True, color=Color(rgb="FFFFFFFF"))
    header_fill = PatternFill(patternType="solid", fgColor=Color(indexed=11))
    header_row = 1

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data rows
    first_data_row = header_row + 1
    for row_offset, entry in enumerate(entries):
        data_row = first_data_row + row_offset

        date_str = ""
        if entry["date"]:
            date_str = format_timesheet_date(entry["date"])

        ws.cell(row=data_row, column=1, value=date_str)
        ws.cell(row=data_row, column=2, value=entry["employee"])
        ws.cell(row=data_row, column=3, value=entry["hours"])
        ws.cell(row=data_row, column=4, value=entry["phase"])
        ws.cell(row=data_row, column=5, value=entry["task"])
        ws.cell(row=data_row, column=6, value=entry["wid"])

    # Add total row
    last_data_row = first_data_row + len(entries) - 1
    total_row = last_data_row + 1

    total_cell = ws.cell(row=total_row, column=1, value="Total")
    total_cell.font = Font(bold=True)

    hours_sum_cell = ws.cell(
        row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})"
    )
    hours_sum_cell.font = Font(bold=True)

    # Set column widths
    column_widths = {
        "A": 30.6640625,
        "B": 4.33203125,
        "C": 4.1640625,
        "D": 3.5,
        "E": 4.0,
        "F": 37.0,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width


def apply_style_fixes(ws) -> None:
    """Apply style fixes to the worksheet."""
    # Insert blank row under logo (at row 2)
    ws.insert_rows(2, 1)

    # Turn off gridlines
    ws.sheet_view.showGridLines = False

    # Add logo to top-left corner
    if LOGO_PATH.exists():
        logo = ExcelImage(str(LOGO_PATH))
        logo.width = 150
        logo.height = 40
        ws.add_image(logo, "A1")

    # Find footer row and set row height
    for row in range(1, ws.max_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and "CCI Design Inc." in cell.value:
                ws.row_dimensions[row].height = 30
                return


def process_project_sheet(
    ws,
    project_id: str,
    hours_dict: dict[tuple[str, str], float],
    invoice_date: date | None = None,
) -> None:
    """Process a single project invoice worksheet."""
    project_name, project_number = parse_project_id(project_id)

    # Calculate rows to delete
    rows_to_delete = calculate_rows_to_delete(hours_dict)

    # Populate invoice data (before deleting rows)
    populate_invoice_sheet(ws, project_name, project_number, hours_dict, invoice_date)

    # Delete empty rows (from bottom to top)
    delete_rows_from_sheet(ws, rows_to_delete)

    # Apply style fixes
    apply_style_fixes(ws)

    # Rebuild formulas with correct cell references
    rebuild_formulas(ws)


def create_invoice_workbook(
    template_path: Path,
    aggregated_data: dict[str, dict[tuple[str, str], float]],
    detail_entries: dict[str, list[dict]],
    invoice_date: date | None = None,
    silent: bool = False,
):
    """Create the output Excel workbook with all project sheets.

    Args:
        template_path: Path to the Excel template file
        aggregated_data: Aggregated hours by project and (task, phase)
        detail_entries: Individual entries grouped by project
        invoice_date: Optional date override for invoices
        silent: If True, suppress print statements (for API usage)
    """
    if not silent:
        print(f"Loading template: {template_path}")

    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    # Load template
    wb = load_workbook(str(template_path))
    template_ws = wb.active
    template_name = template_ws.title

    # Sort projects alphabetically by name
    sorted_projects = sorted(
        aggregated_data.keys(),
        key=lambda p: p.split(":")[0].strip().lower(),
    )

    # Process each project
    for project_id in sorted_projects:
        hours = dict(aggregated_data[project_id])
        entries = detail_entries[project_id]

        phase_totals = calculate_phase_hours(hours)
        total_hours = sum(phase_totals.values())
        active_phases = [p for p in PHASE_ORDER if phase_totals.get(p, 0) > 0]
        removed_phases = [p for p in PHASE_ORDER if phase_totals.get(p, 0) == 0]

        if not silent:
            print(f"\nProcessing: {project_id}")
            print(f"  - {len(entries)} entries, {total_hours:.1f} total hours")
            if active_phases:
                phase_str = ", ".join(f"{p} ({phase_totals[p]:.1f}h)" for p in active_phases)
                print(f"  - Phases: {phase_str}")
            if removed_phases:
                print(f"  - Removing empty phases: {', '.join(removed_phases)}")

        # Create invoice sheet (copy from template)
        invoice_ws = wb.copy_worksheet(template_ws)
        invoice_ws.title = make_sheet_name(project_id, " A")
        process_project_sheet(invoice_ws, project_id, hours, invoice_date)

        # Create timesheet sheet
        timesheet_name = make_sheet_name(project_id, " B")
        create_timesheet_sheet(wb, timesheet_name, entries)

    # Remove the original template sheet
    del wb[template_name]

    return wb


# =============================================================================
# MAIN ENTRY POINTS
# =============================================================================


def _process_invoices(
    input_file: Path,
    invoice_date: date | None = None,
    silent: bool = False,
) -> InvoiceResult:
    """
    Core invoice processing logic shared by file and bytes generators.

    Args:
        input_file: Path to the monthly report Numbers file
        invoice_date: Optional date override for invoices
        silent: If True, suppress print statements (for API usage)

    Returns:
        InvoiceResult with workbook and metadata

    Raises:
        FileNotFoundError: Input file doesn't exist
        ValueError: Validation errors in input data
    """
    # Read input data
    all_entries = read_timesheet_data(input_file)
    unique_projects = set(e["project_id"] for e in all_entries)
    if not silent:
        print(f"Found {len(all_entries)} entries across {len(unique_projects)} unique Project IDs")

    # Filter non-projects
    filtered_entries, excluded_count = filter_non_projects(all_entries)
    if excluded_count > 0 and not silent:
        print(f"Filtering non-projects... {excluded_count} excluded (Office, Vacation, Holiday, etc.)")

    # Filter zero hours
    filtered_entries = filter_zero_hours(filtered_entries)

    # Validate project IDs
    project_errors = validate_project_ids(filtered_entries)
    if project_errors:
        if not silent:
            print("\nERROR: Invalid Project IDs found:")
            for error in project_errors:
                print(f"  - {error}")
        raise ValueError("\n".join(project_errors))

    # Validate codes
    code_errors = validate_codes(filtered_entries)
    if code_errors:
        if not silent:
            print("\nERROR: Invalid Task/Phase codes found:")
            for error in code_errors:
                print(f"  - {error}")
        raise ValueError("\n".join(code_errors))

    if not silent:
        print("Validating codes... OK")

    # Check for billable projects
    if not filtered_entries:
        raise ValueError("No billable projects found in input file")

    # Aggregate and group
    aggregated = aggregate_hours(filtered_entries)
    grouped = group_by_project(filtered_entries)

    if not silent:
        print(f"{len(aggregated)} billable projects to process")

    # Create invoice workbook
    wb = create_invoice_workbook(TEMPLATE_PATH, aggregated, grouped, invoice_date, silent)

    # Calculate totals
    total_hours = sum(sum(hours.values()) for hours in aggregated.values())

    return InvoiceResult(
        workbook=wb,
        project_count=len(aggregated),
        total_hours=total_hours,
        entries=filtered_entries,
    )


def _generate_filename_from_entries(entries: list[dict]) -> str:
    """Generate output filename in YYYY_MM format without version suffix."""
    for entry in entries:
        if entry["date"]:
            year = entry["date"].year
            month = entry["date"].month
            return f"invoices_{year}_{month:02d}.xlsx"

    # Fallback to current date
    today = date.today()
    return f"invoices_{today.year}_{today.month:02d}.xlsx"


def generate_invoices_to_bytes(
    input_file: Path,
    invoice_date: date | None = None,
) -> tuple[bytes, str, int, float]:
    """
    Generate invoices and return as bytes (for API usage).

    Args:
        input_file: Path to the monthly report Numbers file
        invoice_date: Optional date override for invoices

    Returns:
        Tuple of (excel_bytes, filename, project_count, total_hours)

    Raises:
        FileNotFoundError: Input file not found
        ValueError: Validation errors (project IDs, codes, no billable projects)
    """
    result = _process_invoices(input_file, invoice_date, silent=True)

    # Generate filename from entries
    output_filename = _generate_filename_from_entries(result.entries)

    # Save to BytesIO instead of disk
    buffer = BytesIO()
    result.workbook.save(buffer)
    buffer.seek(0)

    return (
        buffer.getvalue(),
        output_filename,
        result.project_count,
        result.total_hours,
    )


def generate_invoices(input_file: Path) -> Path:
    """
    Main entry point for invoice generation (file output).

    Args:
        input_file: Path to the monthly report Numbers file

    Returns:
        Path to the generated invoice file
    """
    result = _process_invoices(input_file, invoice_date=None, silent=False)

    # Generate versioned output path
    output_file = generate_output_filename(input_file, result.entries)
    print(f"\nSorting sheets alphabetically...")
    print(f"Writing output: {output_file}")

    result.workbook.save(str(output_file))

    print(f"\nComplete! Generated invoices for {result.project_count} projects ({result.total_hours:.1f} total hours)")
    print(f"Output: {output_file}")

    return output_file
